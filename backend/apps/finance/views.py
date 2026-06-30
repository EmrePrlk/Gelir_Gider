import csv
import io
import json
import logging
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Category, Transaction, UserExcelSchema
from .serializers import (
    CategorySerializer, TransactionSerializer, UserExcelSchemaSerializer,
)

logger = logging.getLogger(__name__)

PARSE_SYSTEM_PROMPT = """
Sen bir banka ekstresi parser'ısın. Sana verilen CSV, PDF veya Excel metnini analiz edip
her işlem satırı için JSON çıktısı üreteceksin.

Kurallar:
- date: YYYY-MM-DD formatına çevir
- amount: Float, pozitif sayı (sign'ı type alanıyla belirt)
- type: "income" veya "expense" (pozitif gelir, negatif/gider harcama)
- category: [market, restoran, ulaşım, fatura, sağlık, eğlence, giyim, teknoloji, kira, maaş, yatırım, diğer] listesinden seç
- description: Orijinal metni koru, gereksiz boşlukları temizle

Sadece JSON array döndür, başka açıklama yok.
Format: [{"date":"2026-01-15","amount":250.00,"description":"...","category":"market","type":"expense"}]

Boş veya başlık satırlarını atla. Sadece gerçek işlem satırlarını döndür.
"""


def _extract_text_from_file(uploaded_file) -> str:
    """Extract text from uploaded CSV, PDF or Excel file."""
    name = uploaded_file.name.lower()

    if name.endswith('.csv'):
        content = uploaded_file.read().decode('utf-8-sig', errors='replace')
        return content

    if name.endswith('.pdf'):
        try:
            from pdfminer.high_level import extract_text
            pdf_bytes = uploaded_file.read()
            text = extract_text(io.BytesIO(pdf_bytes))
            return text
        except Exception as e:
            logger.error("PDF extraction failed: %s", e)
            raise ValueError("PDF dosyası okunamadı")

    if name.endswith('.xlsx') or name.endswith('.xls'):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), read_only=True, data_only=True)
            ws = wb.active
            lines = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else '' for c in row]
                if any(c.strip() for c in cells):
                    lines.append('\t'.join(cells))
            wb.close()
            return '\n'.join(lines)
        except Exception as e:
            logger.error("Excel extraction failed: %s", e)
            raise ValueError("Excel dosyası okunamadı")

    raise ValueError("Sadece CSV, PDF veya Excel (.xlsx) dosyaları desteklenir")


def _call_anthropic(text: str) -> list[dict]:
    """Send text to Anthropic API and return parsed transaction list."""
    from apps.core.ai import call_anthropic

    raw = call_anthropic(PARSE_SYSTEM_PROMPT, text)
    if not raw:
        return _fallback_csv_parse(text)

    try:
        if raw.startswith('```'):
            raw = raw.split('```')[1]
            if raw.startswith('json'):
                raw = raw[4:]
        return json.loads(raw)
    except Exception as e:
        logger.error("JSON parse failed: %s", e)
        return _fallback_csv_parse(text)


def _fallback_csv_parse(text: str) -> list[dict]:
    """Basic CSV parse when Anthropic is unavailable."""
    rows = []
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        try:
            amount_str = next(
                (v for k, v in row.items() if 'tutar' in k.lower() or 'amount' in k.lower() or 'miktar' in k.lower()),
                None,
            )
            date_str = next(
                (v for k, v in row.items() if 'tarih' in k.lower() or 'date' in k.lower()),
                None,
            )
            desc_str = next(
                (v for k, v in row.items() if 'açıkla' in k.lower() or 'description' in k.lower() or 'detay' in k.lower()),
                None,
            ) or next(iter(row.values()), '')

            if not amount_str or not date_str:
                continue

            amount_clean = amount_str.replace('.', '').replace(',', '.').replace(' ', '').replace('₺', '').replace('TL', '')
            amount = float(amount_clean)
            tx_type = 'expense' if amount < 0 else 'income'
            amount = abs(amount)

            rows.append({
                'date': date_str.strip(),
                'amount': amount,
                'description': desc_str.strip() if desc_str else '',
                'category': 'diğer',
                'type': tx_type,
            })
        except (ValueError, StopIteration):
            continue
    return rows


# ─── ViewSets ───────────────────────────────────────────────────────────────

class CategoryViewSet(viewsets.ModelViewSet):
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = Category.objects.filter(user=self.request.user)
        type_param = self.request.query_params.get('type')
        if type_param:
            qs = qs.filter(type=type_param)
        return qs.order_by('name')

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class TransactionViewSet(viewsets.ModelViewSet):
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = Transaction.objects.filter(user=self.request.user).select_related('category')

        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        category_id = self.request.query_params.get('category')
        tx_type = self.request.query_params.get('type')
        search = self.request.query_params.get('search')

        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if category_id:
            qs = qs.filter(category_id=category_id)
        if tx_type:
            qs = qs.filter(type=tx_type)
        if search:
            qs = qs.filter(description__icontains=search)

        return qs

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


# ─── Summary ────────────────────────────────────────────────────────────────

class SummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = date.today()
        month_start = today.replace(day=1)

        base = Transaction.objects.filter(user=request.user)

        current = base.filter(date__gte=month_start)
        income = current.filter(type='income').aggregate(s=Sum('amount'))['s'] or Decimal('0')
        expense = current.filter(type='expense').aggregate(s=Sum('amount'))['s'] or Decimal('0')

        # Last 12 months trend — single query grouped by year/month/type
        start_month = today.month - 11
        start_year = today.year
        while start_month <= 0:
            start_month += 12
            start_year -= 1
        trend_start = date(start_year, start_month, 1)

        monthly_agg = (
            base.filter(date__gte=trend_start)
            .values('date__year', 'date__month', 'type')
            .annotate(total=Sum('amount'))
        )
        monthly_lookup: dict = {}
        for row in monthly_agg:
            key = (row['date__year'], row['date__month'])
            if key not in monthly_lookup:
                monthly_lookup[key] = {'income': Decimal('0'), 'expense': Decimal('0')}
            if row['type'] in ('income', 'expense'):
                monthly_lookup[key][row['type']] = row['total']

        trend = []
        for i in range(11, -1, -1):
            month = today.month - i
            year = today.year
            while month <= 0:
                month += 12
                year -= 1
            data = monthly_lookup.get((year, month), {})
            trend.append({
                'month': f"{year}-{month:02d}",
                'income': float(data.get('income', Decimal('0'))),
                'expense': float(data.get('expense', Decimal('0'))),
            })

        return Response({
            'income': float(income),
            'expense': float(expense),
            'net': float(income - expense),
            'monthly_trend': trend,
        })


# ─── Category Breakdown ─────────────────────────────────────────────────────

class CategoryBreakdownView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = date.today()
        date_from = request.query_params.get('date_from', str(today.replace(day=1)))
        date_to = request.query_params.get('date_to', str(today))
        tx_type = request.query_params.get('type', 'expense')

        qs = Transaction.objects.filter(
            user=request.user,
            type=tx_type,
            date__gte=date_from,
            date__lte=date_to,
        ).select_related('category')

        totals: dict = {}
        grand_total = Decimal('0')

        for tx in qs:
            if tx.category:
                key = tx.category_id
                if key not in totals:
                    totals[key] = {
                        'category_id': tx.category_id,
                        'category_name': tx.category.name,
                        'category_color': tx.category.color,
                        'category_icon': tx.category.icon,
                        'total': Decimal('0'),
                    }
                totals[key]['total'] += tx.amount
                grand_total += tx.amount
            else:
                # Uncategorized bucket
                if 'uncategorized' not in totals:
                    totals['uncategorized'] = {
                        'category_id': None,
                        'category_name': 'Kategorisiz',
                        'category_color': '#4A4A62',
                        'category_icon': 'HelpCircle',
                        'total': Decimal('0'),
                    }
                totals['uncategorized']['total'] += tx.amount
                grand_total += tx.amount

        result = []
        for item in sorted(totals.values(), key=lambda x: x['total'], reverse=True):
            pct = float(item['total'] / grand_total * 100) if grand_total else 0
            result.append({**item, 'total': float(item['total']), 'percentage': round(pct, 1)})

        return Response(result)


# ─── Excel Schema ───────────────────────────────────────────────────────────

class ExcelSchemaView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            schema = UserExcelSchema.objects.get(user=request.user)
            return Response(UserExcelSchemaSerializer(schema).data)
        except UserExcelSchema.DoesNotExist:
            return Response(None)

    def put(self, request):
        schema, _ = UserExcelSchema.objects.get_or_create(user=request.user)
        serializer = UserExcelSchemaSerializer(schema, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


# ─── Import ─────────────────────────────────────────────────────────────────

class ImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'detail': 'Dosya gerekli'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            text = _extract_text_from_file(uploaded)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        rows = _call_anthropic(text)

        # Normalize rows
        normalized = []
        for row in rows:
            try:
                normalized.append({
                    'date': str(row.get('date', '')),
                    'amount': abs(float(row.get('amount', 0))),
                    'description': str(row.get('description', '')),
                    'category': str(row.get('category', '')),
                    'type': row.get('type', 'expense'),
                })
            except (ValueError, TypeError):
                continue

        return Response({'preview': normalized, 'count': len(normalized)})


class ImportConfirmView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        rows = request.data.get('rows', [])
        if not rows:
            return Response({'detail': 'Satır gerekli'}, status=status.HTTP_400_BAD_REQUEST)

        # Build category lookup
        user_categories = {
            c.name.lower(): c
            for c in Category.objects.filter(user=request.user)
        }

        created = []
        for row in rows:
            category = user_categories.get(str(row.get('category', '')).lower())
            try:
                amount = Decimal(str(abs(float(row.get('amount', 0))))).quantize(Decimal('0.01'))
                tx_date = row.get('date') or str(date.today())
                tx = Transaction.objects.create(
                    user=request.user,
                    date=tx_date,
                    amount=amount,
                    description=str(row.get('description', '')),
                    category=category,
                    type=row.get('type', 'expense'),
                    source='import',
                )
                created.append(tx.id)
            except (InvalidOperation, Exception) as e:
                logger.warning("Skipping import row: %s — %s", row, e)
                continue

        return Response({'created': len(created), 'ids': created}, status=status.HTTP_201_CREATED)


# ─── Export ─────────────────────────────────────────────────────────────────

class ExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        qs = Transaction.objects.filter(user=request.user).select_related('category').order_by('date')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'İşlemler'
        ws.append(['Tarih', 'Açıklama', 'Kategori', 'Tür', 'Tutar', 'Notlar'])

        for tx in qs:
            ws.append([
                str(tx.date),
                tx.description,
                tx.category.name if tx.category else '',
                'Gelir' if tx.type == 'income' else 'Gider',
                float(tx.amount),
                tx.notes,
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        res = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        res['Content-Disposition'] = 'attachment; filename="islemler.xlsx"'
        return res
